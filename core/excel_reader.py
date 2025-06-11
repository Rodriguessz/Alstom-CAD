from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from typing import Optional

class ExcelReader:
    """A classe ExcelReader oferece uma interface para interagir com arquivos do Excel,
    permitindo carregar e manipular dados em planilhas de forma simples e eficiente.
    
    Responsabilidades Principais:
    - Verificação da Existência do Arquivo: Verifica se o arquivo Excel especificado 
      existe no sistema de arquivos antes de tentar carregá-lo.
    - Abertura de Planilhas: Carrega uma planilha do arquivo Excel e permite a seleção 
      da aba desejada, com suporte para selecionar a primeira aba por padrão.
    - Busca de Valores: Permite procurar células baseadas em seus valores, retornando a
      célula correspondente se encontrada, ou a busca por células em posições específicas.
    - Atualização de Células: Facilita a atualização de células encontradas, seja 
      por valor, seja por posição.
    - Aplicação de Modificações: Permite aplicar modificações em relação a estados
      detectados em um documento, atualizando as células correspondentes de acordo.
    - Salvamento de Alterações: Salva as alterações realizadas na planilha em um novo 
      local ou sobrescreve o arquivo original.

    A classe utiliza a biblioteca OpenPyXL para manipulação de arquivos .xlsx, proporcionando
    uma maneira eficiente de trabalhar com dados tabulares, sendo especialmente útil em 
    automações e análises que requerem interação com documentos do Excel."""

    def __init__(self, file_path: Path ):
        self.file_path = file_path;
        self.workbook = None
        self.sheet: Optional[Worksheet] = None
        self.columns = ["A", "B"]

    def exists(self) -> bool:
        """Verifica se  o caminho do arquivo selecionado pelo usuário realmente existe"""
        return self.file_path.exists()
    
    def open_excel_sheet(self, sheet_name: str = None):
        """Abre a planilha e carrega a aba especificada pelo usuário (ou a primeira, por padrão)"""

        if not self.exists():
            raise FileNotFoundError(f"Arquivo Excel não encontrado: {self.file_path}")

        self.workbook = load_workbook(self.file_path)

        if sheet_name:
            if sheet_name in self.workbook.sheetnames:
                self.sheet = self.workbook[sheet_name]
            else:
                raise ValueError(f"Aba '{sheet_name}' não encontrada no arquivo.")
        else:
            # Se nenhum nome for passado, seleciona a primeira aba
            self.sheet = self.workbook.active

    def find_cell_by_value(self, value: str):
        """Procura por um valor na planilha inteira e retorna a célula (se existir)"""
        if not self.sheet:
            raise ValueError("Aba Excel não carregada.")

        #Percore as linhas da planilha
        for row in self.sheet.iter_rows():
            #Percorre as celulas presentes na linha atual e compara com o valor passado;
            for cell in row:
                #Padroniza os valores removendo espaçoes e deixando em uppercase para comparação
                if str(cell.value).strip().lower() == str(value).strip().lower():
                    return cell
        return None
    
    def update_cell_by_value(self, target_value: str, new_value : str):
        """Procura uma célula pelo valor e atualiza com novo valor"""
        cell = self.find_cell_by_value(target_value)
        if cell:
            cell.value = str(new_value).upper();
            return True
        return False
    
    def find_cell_by_position(self, position: str):
        """Procura na planilha inteira e retorna a célula na posição indicada (se existir)"""
        if not self.sheet:
            raise ValueError("Aba Excel não carregada.")
        try:
            cell = self.sheet[position]
            return cell
        except KeyError:
            raise ValueError(f"A célula {position} não existe na planilha.")

    def update_cell_by_position(self, position: str, new_value):
        """Procura uma célula pela posição indicada e atualiza com um novo valor."""
        cell = self.find_cell_by_position(position);

        if cell:
            cell.value = str(new_value).upper();
            return True
        
        return False;
    
    def aplly_relays_modifications(self, modifications: list):
        """Aplica todas as modificações detecadas no desenho na planilha."""
        
        for mod in modifications:

            if "sin" in mod:        
                r_cell = self.find_cell_by_value(mod["sin"]["r"]);

                if r_cell:
                
                    r_cell_column = r_cell.column_letter;
                
                    r_cell_row = r_cell.row;
                    
                    r_cell_column_index = column_index_from_string(r_cell_column);

                    sin_cell_column_index = r_cell_column_index + 1;
                    sin_cell_column = get_column_letter(sin_cell_column_index);

                    sin_cell_position = f"{sin_cell_column}{r_cell_row}"
                

                    self.update_cell_by_position(sin_cell_position, str(mod["sin"]["new_state"]));  
            

            if "r" in mod:
                self.update_cell_by_value(mod["r"]["previous_state"], str(mod["r"]["new_state"]));

    def save_changes(self, new_path: Optional[Path] = None):
        """Salva as alterações feitas na planilha em um novo caminho ou sobreescreve o arquivo original."""
        
        if not self.workbook:
            raise ValueError("Workbook não carregado.")

        #Salva em um novo local ou no local padrão;
        path_to_save = new_path or self.file_path
        self.workbook.save(path_to_save)

    